import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string


#ruta =  "C:\\Users\\cromero\\Desktop\\Nueva carpeta (9)\\supermarket_sales.xlsx"

def funcion_reporte(ruta):
    """input sales_mes.xlsx / output report_sales.xlsx"""

    archivo_excel = pd.read_excel(ruta)
    tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    mes_extension = ruta.split('_')[1]
    tabla_pivote.to_excel(f'sales_{mes_extension}', startrow=4, sheet_name='Report')

    wb = load_workbook(f'sales_{mes_extension}')
    pestaña = wb['Report']

    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row


    # Grafico


    barchart = BarChart()

    data = Reference(pestaña, min_col=min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila)
    categorias = Reference(pestaña, min_col=min_col,max_col=min_col, min_row=min_fila+1, max_row=max_fila)

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categorias)

    pestaña.add_chart(barchart, 'b12')
    barchart.title = 'VENTAS'
    barchart.style = 2

    abecedario = (string.ascii_uppercase)

    abecedario_excel = abecedario[0:max_col]

    for i in abecedario_excel:
        if i != 'A':
            pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
            pestaña[f'{i}{max_fila+1}'].style = 'Currency'

    pestaña[f'{abecedario_excel[0]}{max_col+1}'] = 'Total'


    pestaña['A1'] = 'Reporte'
    mes = mes_extension.split('.')[0]
    pestaña['A2'] = mes

    pestaña['A1'].font = Font('Arial', bold=True, size=20)
    pestaña['A2'].font = Font('Arial', bold=True, size=12)

    wb.save(f'sales_{mes_extension}')

    return 

   
funcion_reporte('C:\\Users\\cromero\\Desktop\\Nueva carpeta (9)\\supermarket_sales.xlsx')
